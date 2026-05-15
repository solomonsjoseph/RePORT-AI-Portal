"""Pairing tests: PDF+dataset matching, collisions, deduplication."""

from pathlib import Path

import openpyxl
import pytest

from scripts.source_truth.study_intake import (
    IntakeManifest,
    Pair,
    ReviewEntry,
    pair_files,
    REASON_UNPAIRED_PDF,
    REASON_UNPAIRED_DATASET,
    REASON_DUPLICATE_SHA,
    REASON_MISMATCHED_SHA,
    REASON_COLLISION,
)


@pytest.fixture
def study_dir(tmp_path: Path) -> Path:
    """Create a study directory structure."""
    study = tmp_path / "test_study"
    (study / "datasets").mkdir(parents=True)
    (study / "annotated_pdfs").mkdir(parents=True)
    return study


def test_simple_pairing(study_dir: Path):
    """One xlsx + one pdf with matching form code → 1 paired, 0 review."""
    # Create dataset
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    wb.save(study_dir / "datasets" / "1A_ICScreening.xlsx")

    # Create PDF
    (study_dir / "annotated_pdfs" / "1A_ICScreening.pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 1
    assert len(manifest.to_review) == 0
    assert manifest.paired[0].form_code == "1a"
    assert manifest.paired[0].dataset_path.stem == "1A_ICScreening"
    assert manifest.paired[0].pdf_path.stem == "1A_ICScreening"


def test_filename_normalization(study_dir: Path):
    """Variants with spaces, underscores, version suffixes still match."""
    # Create dataset with variant naming
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    wb.save(study_dir / "datasets" / "1A_ic_screening_v2.xlsx")

    # Create PDF with different variant
    (study_dir / "annotated_pdfs" / "1A ICScreening (1).pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 1
    assert len(manifest.to_review) == 0


def test_form_code_prefix_match(study_dir: Path):
    """Same form-code prefix matches; different prefixes do not."""
    # Pair 1A
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    wb.save(study_dir / "datasets" / "1A_foo.xlsx")
    (study_dir / "annotated_pdfs" / "1A_bar.pdf").write_text("PDF content")

    # Unpaired 2B dataset
    wb2 = openpyxl.Workbook()
    wb2.active.append(["SUBJID", "AGE"])
    wb2.save(study_dir / "datasets" / "2B_x.xlsx")

    # Unpaired 2C PDF
    (study_dir / "annotated_pdfs" / "2C_x.pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 1
    assert manifest.paired[0].form_code == "1a"
    # 2B_x and 2C_x should be unpaired
    assert len(manifest.to_review) == 2
    reasons = {e.reason for e in manifest.to_review}
    assert REASON_UNPAIRED_DATASET in reasons
    assert REASON_UNPAIRED_PDF in reasons


def test_unpaired_pdf(study_dir: Path):
    """PDF without matching dataset → review with unpaired_pdf."""
    (study_dir / "annotated_pdfs" / "1A_orphan.pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 0
    assert len(manifest.to_review) == 1
    assert manifest.to_review[0].reason == REASON_UNPAIRED_PDF
    assert "1A_orphan" in manifest.to_review[0].label


def test_unpaired_dataset(study_dir: Path):
    """Dataset without matching PDF → review with unpaired_dataset."""
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    wb.save(study_dir / "datasets" / "1A_orphan.xlsx")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 0
    assert len(manifest.to_review) == 1
    assert manifest.to_review[0].reason == REASON_UNPAIRED_DATASET
    assert "1A_orphan" in manifest.to_review[0].label


def test_duplicate_sha_dedupes(study_dir: Path):
    """Two xlsx files with identical content and same form-code → 1 paired, 0 review."""
    # Create the same dataset twice
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    content = openpyxl.utils.get_column_letter(1)  # dummy content reference
    wb.save(study_dir / "datasets" / "1A_data_v1.xlsx")
    wb.save(study_dir / "datasets" / "1A_data_v2.xlsx")

    # Create PDF
    (study_dir / "annotated_pdfs" / "1A_form.pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    # The two xlsx files have identical SHA, so deduplication keeps one
    assert len(manifest.paired) == 1
    assert len(manifest.to_review) == 0


def test_mismatched_sha_to_review(study_dir: Path):
    """Two xlsx files with different content and same form-code → routed to review.

    When _dedup_or_flag detects mismatched SHA, it:
    1. Routes both xlsx files with REASON_MISMATCHED_SHA
    2. Returns empty list []
    3. The collision check then routes the PDF with REASON_COLLISION
    Total: 3 entries in to_review (2 xlsx + 1 pdf)
    """
    # Create two different datasets
    wb1 = openpyxl.Workbook()
    wb1.active.append(["SUBJID", "AGE"])
    wb1.save(study_dir / "datasets" / "1A_data_v1.xlsx")

    wb2 = openpyxl.Workbook()
    wb2.active.append(["SUBJID", "AGE", "EXTRA"])
    wb2.save(study_dir / "datasets" / "1A_data_v2.xlsx")

    # Create PDF
    (study_dir / "annotated_pdfs" / "1A_form.pdf").write_text("PDF content")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 0
    assert len(manifest.to_review) == 3  # 2 xlsx (mismatched_sha) + 1 pdf (collision)
    # Both xlsx files should have mismatched_sha
    mismatched = [e for e in manifest.to_review if e.reason == REASON_MISMATCHED_SHA]
    assert len(mismatched) == 2
    # PDF should have collision
    collisions = [e for e in manifest.to_review if e.reason == REASON_COLLISION]
    assert len(collisions) == 1


def test_collision(study_dir: Path):
    """Two PDFs with same form-code prefix → review with mixed reasons.

    When two PDFs have the same prefix with mismatched SHA:
    1. _dedup_or_flag on PDFs routes both as mismatched_sha (returns [])
    2. Collision check sees len(pdfs) != 1, routes dataset as collision
    Total: 3 entries (2 mismatched_sha pdfs + 1 collision dataset)
    """
    wb = openpyxl.Workbook()
    wb.active.append(["SUBJID", "AGE"])
    wb.save(study_dir / "datasets" / "1A_form.xlsx")

    # Create two different PDFs with the same prefix
    (study_dir / "annotated_pdfs" / "1A_page1.pdf").write_text("PDF content 1")
    (study_dir / "annotated_pdfs" / "1A_page2.pdf").write_text("PDF content 2")

    manifest = pair_files(study_dir)
    assert len(manifest.paired) == 0
    assert len(manifest.to_review) == 3
    mismatches = [e for e in manifest.to_review if e.reason == REASON_MISMATCHED_SHA]
    collisions = [e for e in manifest.to_review if e.reason == REASON_COLLISION]
    assert len(mismatches) == 2  # Two PDFs with different content
    assert len(collisions) == 1  # Dataset routed as collision
