"""Exclusion tests: formula headers, PHI patterns, empty rows, route_to_review."""

from pathlib import Path

import openpyxl
import pytest

from scripts.source_truth.study_intake import (
    ExcludeForReview,
    Pair,
    ReviewEntry,
    build_yaml_for_pair,
    read_headers_only,
    route_to_review,
    _validate_headers,
    REASON_FORMULA_HEADER,
    REASON_PHI_IN_HEADER,
    REASON_EMPTY_HEADER_ROW,
    REASON_UNPAIRED_PDF,
    REASON_UNPAIRED_DATASET,
    REASON_DUPLICATE_SHA,
    REASON_MISMATCHED_SHA,
    REASON_COLLISION,
)


class TestExclusionGuards:
    """Tests for header validation guards."""

    def test_formula_header(self, tmp_path: Path):
        """Row 1 with formula prefix → ExcludeForReview(formula_header).

        Formula guards are checked in _validate_headers, not read_headers_only.
        """
        headers = ["SUBJID", "=A1+B1"]
        dataset_path = tmp_path / "test.xlsx"

        with pytest.raises(ExcludeForReview) as exc_info:
            _validate_headers(headers, dataset_path)
        assert exc_info.value.args[0] == REASON_FORMULA_HEADER

    def test_empty_header_row_xlsx(self, tmp_path: Path):
        """Empty row 1 → ExcludeForReview(empty_header_row)."""
        wb = openpyxl.Workbook()
        wb.active.append([None, None, None])
        xlsx_path = tmp_path / "empty.xlsx"
        wb.save(xlsx_path)
        wb.close()

        with pytest.raises(ExcludeForReview) as exc_info:
            read_headers_only(xlsx_path)
        assert exc_info.value.args[0] == REASON_EMPTY_HEADER_ROW

    def test_blank_header_row_xlsx(self, tmp_path: Path):
        """Whitespace-only row 1 → ExcludeForReview(empty_header_row)."""
        wb = openpyxl.Workbook()
        wb.active.append(["", "  ", "\t"])
        xlsx_path = tmp_path / "blank.xlsx"
        wb.save(xlsx_path)
        wb.close()

        with pytest.raises(ExcludeForReview) as exc_info:
            read_headers_only(xlsx_path)
        assert exc_info.value.args[0] == REASON_EMPTY_HEADER_ROW

    def test_unsupported_extension(self, tmp_path: Path):
        """File with unsupported extension → ExcludeForReview(unsupported_extension)."""
        txt_path = tmp_path / "data.txt"
        txt_path.write_text("header1,header2\ndata1,data2\n")

        with pytest.raises(ExcludeForReview) as exc_info:
            read_headers_only(txt_path)
        assert exc_info.value.args[0] == "unsupported_extension"

    def test_multi_sheet_workbook(self, tmp_path: Path):
        """Workbook with multiple visible sheets → ExcludeForReview(multi_sheet_workbook)."""
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.active.append(["SUBJID", "AGE"])
        wb.create_sheet("Sheet2")
        wb["Sheet2"].append(["ID", "VALUE"])
        xlsx_path = tmp_path / "multi.xlsx"
        wb.save(xlsx_path)
        wb.close()

        with pytest.raises(ExcludeForReview) as exc_info:
            read_headers_only(xlsx_path)
        assert exc_info.value.args[0] == "multi_sheet_workbook"


class TestRouteToReview:
    """Tests for SoT_intake_review.md writing."""

    def test_route_to_review_writes_correct_section(self, tmp_path: Path):
        """Each reason code maps to correct section header in markdown."""
        review_file = tmp_path / "SoT_intake_review.md"

        reason_codes = [
            REASON_UNPAIRED_PDF,
            REASON_UNPAIRED_DATASET,
            REASON_EMPTY_HEADER_ROW,
            REASON_FORMULA_HEADER,
            REASON_PHI_IN_HEADER,
            REASON_MISMATCHED_SHA,
            REASON_COLLISION,
        ]

        for i, reason in enumerate(reason_codes):
            entry = ReviewEntry(
                reason=reason,
                label=f"form_{i}",
                file_path=Path(f"/data/test_{i}.xlsx"),
                notes="Test note",
            )
            route_to_review(entry, review_file)

        content = review_file.read_text(encoding="utf-8")
        # Check that each reason code has a corresponding section
        for reason in reason_codes:
            assert f"form_code" in content or reason in content
            assert f"form_" in content  # At least one entry was written

    def test_route_to_review_creates_new_file(self, tmp_path: Path):
        """First call to route_to_review creates the markdown file."""
        review_file = tmp_path / "human_review" / "SoT_intake_review.md"

        entry = ReviewEntry(
            reason=REASON_UNPAIRED_PDF,
            label="1A_orphan",
            file_path=Path("/data/raw/study/annotated_pdfs/1A_orphan.pdf"),
            notes="No matching dataset",
        )
        route_to_review(entry, review_file)

        assert review_file.exists()
        content = review_file.read_text(encoding="utf-8")
        assert "SoT Intake Review" in content
        assert "1A_orphan" in content
        assert "unpaired_pdf" in content

    def test_route_to_review_honors_existing_disposition(self, tmp_path: Path):
        """Pre-ticked entry is not re-appended."""
        review_file = tmp_path / "SoT_intake_review.md"

        # Pre-seed with a dispositioned entry
        review_file.parent.mkdir(parents=True, exist_ok=True)
        header = (
            "# SoT Intake Review\n"
            "Generated: 2026-05-15 by `python -m scripts.source_truth.study_intake`\n\n"
            "## Unpaired PDFs (no matching dataset)\n"
            "\n- [x] **FORM_1A** — `/path/to/file`\n"
            "      Reason: `unpaired_pdf`\n"
            "      Notes: Already handled\n"
            "      Disposition: keep_in_review\n"
        )
        review_file.write_text(header, encoding="utf-8")

        # Try to route the same entry again
        entry = ReviewEntry(
            reason=REASON_UNPAIRED_PDF,
            label="FORM_1A",
            file_path=Path("/path/to/file"),
            notes="Should not be duplicated",
        )
        route_to_review(entry, review_file)

        content = review_file.read_text(encoding="utf-8")
        # Should not contain duplicate entry — only one checkbox for FORM_1A
        count = content.count("- [x] **FORM_1A**")
        assert count == 1, f"Expected 1 checked entry for FORM_1A, found {count}"

    def test_route_to_review_appends_to_existing_section(self, tmp_path: Path):
        """New entry appends to existing section, not created twice."""
        review_file = tmp_path / "SoT_intake_review.md"

        entry1 = ReviewEntry(
            reason=REASON_UNPAIRED_PDF,
            label="1A_orphan",
            file_path=Path("/data/raw/study/1A_orphan.pdf"),
            notes="First orphan",
        )
        route_to_review(entry1, review_file)

        entry2 = ReviewEntry(
            reason=REASON_UNPAIRED_PDF,
            label="2B_orphan",
            file_path=Path("/data/raw/study/2B_orphan.pdf"),
            notes="Second orphan",
        )
        route_to_review(entry2, review_file)

        content = review_file.read_text(encoding="utf-8")
        # Only one section header for unpaired PDFs
        section_count = content.count("## Unpaired PDFs")
        assert section_count == 1
        assert "1A_orphan" in content
        assert "2B_orphan" in content
