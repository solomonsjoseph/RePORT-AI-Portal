"""Tests for scripts/extraction/dataset_pipeline.py — dataset extraction pipeline."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pytest

import config
from scripts.extraction.dataset_pipeline import (
    clean_record_for_json,
    discover_dataset_files,
    extract_datasets,
    extract_single_dataset,
    is_dataframe_empty,
)


class TestCleanRecordForJson:
    def test_converts_nat_to_none(self) -> None:
        record: dict[str, Any] = {"date": pd.NaT, "value": 42}
        result = clean_record_for_json(record)
        assert result["date"] is None
        assert result["value"] == 42

    def test_converts_nan_to_none(self) -> None:
        record: dict[str, Any] = {"score": float("nan"), "name": "test"}
        result = clean_record_for_json(record)
        assert result["score"] is None

    def test_converts_timestamp_to_iso(self) -> None:
        record: dict[str, Any] = {"date": pd.Timestamp("2014-07-28")}
        result = clean_record_for_json(record)
        assert isinstance(result["date"], str)
        assert "2014" in result["date"]

    def test_preserves_normal_values(self) -> None:
        record: dict[str, Any] = {"a": 1, "b": "text", "c": True}
        result = clean_record_for_json(record)
        assert result == {"a": 1, "b": "text", "c": True}

    def test_empty_record(self) -> None:
        assert clean_record_for_json({}) == {}


class TestDiscoverDatasetFiles:
    def test_finds_excel_and_csv(self, tmp_path: Path) -> None:
        (tmp_path / "data.xlsx").write_bytes(b"fake")
        (tmp_path / "info.csv").write_bytes(b"fake")
        (tmp_path / "notes.txt").write_bytes(b"ignored")
        result = discover_dataset_files(tmp_path)
        assert len(result) == 2

    def test_skips_lock_files(self, tmp_path: Path) -> None:
        (tmp_path / "~$locked.xlsx").write_bytes(b"fake")
        (tmp_path / "real.xlsx").write_bytes(b"fake")
        result = discover_dataset_files(tmp_path)
        assert len(result) == 1

    def test_empty_directory_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="No supported"):
            discover_dataset_files(tmp_path)

    def test_missing_directory_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            discover_dataset_files(tmp_path / "nonexistent")

    def test_returns_paths(self, tmp_path: Path) -> None:
        (tmp_path / "data.xlsx").write_bytes(b"fake")
        result = discover_dataset_files(tmp_path)
        assert all(isinstance(p, Path) for p in result)


class TestIsDataframeEmpty:
    def test_empty_dataframe(self) -> None:
        assert is_dataframe_empty(pd.DataFrame())

    def test_non_empty_dataframe(self) -> None:
        assert not is_dataframe_empty(pd.DataFrame({"a": [1]}))

    def test_columns_but_no_rows(self) -> None:
        df = pd.DataFrame(columns=["a", "b"])
        # is_dataframe_empty returns True only when BOTH rows AND columns are 0
        assert not is_dataframe_empty(df)


# ═══════════════════════════════════════════════════════════════════════════
# Staging workspace + drop-event propagation (Task 2)
# ═══════════════════════════════════════════════════════════════════════════


def _make_multisheet_workbook_with_duplicates(path: Path) -> None:
    """Write a two-sheet .xlsx where each sheet has an identical-column pair.

    Sheet ``S1``:
        SUBJID | AGE | SUBJID2   (SUBJID2 duplicates SUBJID → drop event)
    Sheet ``S2``:
        NAME   | VAL | NAME_1    (NAME_1 duplicates NAME    → drop event)
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "S1"
    ws1.append(["SUBJID", "AGE", "SUBJID2"])
    for i in range(3):
        sid = f"SUBJ-{i:03d}"
        ws1.append([sid, 20 + i, sid])  # SUBJID2 identical to SUBJID

    ws2 = wb.create_sheet("S2")
    ws2.append(["NAME", "VAL", "NAME_1"])
    for i in range(3):
        name = f"name_{i}"
        ws2.append([name, i * 10, name])  # NAME_1 identical to NAME

    wb.save(path)


class TestExtractSingleDatasetReturnsDropEvents:
    """Step 1: extract_single_dataset returns a 4-tuple with drop events."""

    def test_returns_four_tuple_with_events(self, tmp_path: Path) -> None:
        src = tmp_path / "dup_multi.xlsx"
        _make_multisheet_workbook_with_duplicates(src)
        out_dir = tmp_path / "out"
        out_dir.mkdir()

        result = extract_single_dataset(
            file_path=src,
            output_dir=out_dir,
            study_name="TEST",
            extraction_ts="2026-04-21T00:00:00+00:00",
        )

        assert isinstance(result, tuple)
        assert len(result) == 4
        ok, count, err, events = result
        assert ok is True
        assert err is None
        assert count == 6  # 3 rows across 2 sheets
        # Events is ALWAYS a list, never None
        assert isinstance(events, list)
        # One drop per sheet → two total
        assert len(events) == 2
        names = {e["name"] for e in events}
        assert names == {"SUBJID2", "NAME_1"}
        # Every event carries file + scope fields (from clean_duplicate_columns)
        for ev in events:
            assert ev["file"] == "dup_multi.xlsx"
            assert ev["scope"] == "dataset-column"

    def test_failure_returns_empty_events_list(self, tmp_path: Path) -> None:
        # Nonexistent / unreadable file → (False, 0, err, [])
        bogus = tmp_path / "does_not_exist.xlsx"
        result = extract_single_dataset(
            file_path=bogus,
            output_dir=tmp_path,
            study_name="TEST",
            extraction_ts="2026-04-21T00:00:00+00:00",
        )
        assert len(result) == 4
        ok, count, err, events = result
        assert ok is False
        assert count == 0
        assert err is not None
        assert events == []


class TestExtractDatasetsDefaultsToStaging:
    """Step 2: extract_datasets defaults output to config.STAGING_DATASETS_DIR."""

    def test_default_output_is_staging(
        self,
        tmp_path: Path,
        monkeypatch_config: Path,  # side-effect: patches config paths to tmp
    ) -> None:
        # Source directory with one real Excel file
        src_dir = tmp_path / "raw_datasets"
        src_dir.mkdir()
        _make_multisheet_workbook_with_duplicates(src_dir / "dup.xlsx")

        # Do NOT pass output_dir → default must resolve to STAGING_DATASETS_DIR
        result = extract_datasets(
            datasets_dir=str(src_dir),
            study_name="TEST",
        )

        assert result["output_dir"] == str(config.STAGING_DATASETS_DIR)
        assert config.STAGING_DATASETS_DIR.exists()
        # JSONL files landed in staging, not in trio_bundle
        jsonls = list(config.STAGING_DATASETS_DIR.glob("*.jsonl"))
        assert jsonls, "expected extracted JSONL under staging"
        assert not list(config.TRIO_DATASETS_DIR.glob("*.jsonl")), (
            "staging rewrite must not write to trio_bundle anymore"
        )


class TestExtractionResultDroppedEvents:
    """Step 3: ExtractionResult.dropped_events aggregates events across files."""

    def test_dropped_events_populated_on_duplicates(
        self,
        tmp_path: Path,
        monkeypatch_config: Path,  # side-effect: patches config paths to tmp
    ) -> None:
        src_dir = tmp_path / "raw_datasets"
        src_dir.mkdir()
        _make_multisheet_workbook_with_duplicates(src_dir / "dup.xlsx")

        result = extract_datasets(
            datasets_dir=str(src_dir),
            study_name="TEST",
        )

        assert "dropped_events" in result
        events = result["dropped_events"]
        assert isinstance(events, list)
        # 2 drop events (one per sheet) from the single input file
        assert len(events) == 2
        names = {e["name"] for e in events}
        assert names == {"SUBJID2", "NAME_1"}

    def test_dropped_events_empty_when_no_duplicates(
        self,
        tmp_path: Path,
        monkeypatch_config: Path,  # side-effect: patches config paths to tmp
    ) -> None:
        # Single-sheet workbook with NO duplicate columns
        src_dir = tmp_path / "raw_datasets"
        src_dir.mkdir()
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["SUBJID", "AGE"])
        for i in range(3):
            ws.append([f"SUBJ-{i}", 20 + i])
        wb.save(src_dir / "clean.xlsx")

        result = extract_datasets(
            datasets_dir=str(src_dir),
            study_name="TEST",
        )

        assert result["dropped_events"] == []

    def test_dropped_events_on_error_path_is_empty_list(
        self,
        tmp_path: Path,
        monkeypatch_config: Path,  # side-effect: patches config paths to tmp
    ) -> None:
        # Missing source dir → error path must still return dropped_events=[]
        missing = tmp_path / "nonexistent"
        result = extract_datasets(
            datasets_dir=str(missing),
            study_name="TEST",
        )
        assert result["dropped_events"] == []
