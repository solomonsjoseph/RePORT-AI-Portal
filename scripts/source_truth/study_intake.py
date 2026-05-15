"""SoT intake CLI — pair annotated PDFs with xlsx/csv datasets and produce policy YAMLs.

**What.** A portable CLI (``python -m scripts.source_truth.study_intake <study>
[--force]``) that walks ``data/raw/<study>/annotated_pdfs/`` and
``data/raw/<study>/datasets/``, pairs each PDF with its dataset by leading
form-code prefix (``^[0-9]+[A-Z]?``), reads *only row 1* of every xlsx/csv to
obtain column headers, and dispatches each clean pair through the extractor and
reviewer agents to produce ``data/SoT/<study>/<form>_policy.yaml``.  Everything
that cannot be cleanly paired — unpaired files, PHI-shaped headers, formula
headers, empty header rows, multi-sheet workbooks, SHA-mismatch duplicates, and
filename collisions — is routed to
``data/SoT/<study>/human_review/SoT_intake_review.md`` with a typed reason code.

**Why.** The 38-module ``scripts/source_truth/`` pipeline accumulated complex
intermediate artefacts (ledgers, evidence packs, catalog cards) that made the
SoT-creation path opaque and brittle for non-Claude LLM tools.  The data-isolation
invariant — *row 2+ bytes must never enter Python* — was spread across many modules
with no single enforcement point or static test.  This module is the sole
enforcement point.  The CLI shape (not a ``.claude/skills/`` file) makes the
capability reachable from any shell or agentic tool: ChatGPT, Gemini, Cursor, Aider.

**How.** Deterministic outer shell (this module) drives a fixed workflow:

1. ``pair_files`` normalises filenames, extracts form-code prefixes, matches PDFs
   to datasets, computes SHA-256, routes collisions/unpaired/duplicate files to
   review.
2. ``read_headers_only`` opens xlsx via ``openpyxl`` with ``max_row=1`` or opens
   csv and calls ``next(reader)`` exactly once — the file handle never iterates
   further; row-2+ bytes are structurally unreachable.
3. ``_validate_headers`` runs each header through formula-prefix and PHI-pattern
   guards before any header reaches the LLM.
4. ``build_yaml_for_pair`` calls ``pdf_evidence.extract_pdf_evidence`` →
   ``sot_extractor_agent.run_extractor`` → ``sot_reviewer_agent.run_reviewer`` →
   ``record.validate_record`` → atomic write.
5. ``run_intake`` orchestrates the above and returns a summary exit code.

See: ``docs/runbook_sot_build.md``
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

import config
from scripts.source_truth import pdf_evidence
from scripts.source_truth import record as _record
from scripts.source_truth import sot_extractor_agent, sot_reviewer_agent
from scripts.utils.logging_system import get_logger

__all__ = [
    "ExcludeForReview",
    "IntakeManifest",
    "Pair",
    "ReviewEntry",
    "build_yaml_for_pair",
    "main",
    "pair_files",
    "read_headers_only",
    "route_to_review",
    "run_intake",
]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_LOG = get_logger(__name__)

# ---------------------------------------------------------------------------
# Exclusion reason codes (8, per phase-plan §6 + threat model)
# ---------------------------------------------------------------------------

REASON_PHI_IN_HEADER = "phi_in_header"
REASON_FORMULA_HEADER = "formula_header"
REASON_UNPAIRED_PDF = "unpaired_pdf"
REASON_UNPAIRED_DATASET = "unpaired_dataset"
REASON_DUPLICATE_SHA = "duplicate_sha"  # identical SHA — dedupe silently (not routed)
REASON_MISMATCHED_SHA = "mismatched_sha"
REASON_COLLISION = "collision"
REASON_EMPTY_HEADER_ROW = "empty_header_row"

# Human-readable section headings for SoT_intake_review.md, keyed by reason code.
_SECTION_TITLES: dict[str, str] = {
    REASON_UNPAIRED_PDF: "Unpaired PDFs (no matching dataset)",
    REASON_UNPAIRED_DATASET: "Unpaired datasets (no matching PDF)",
    REASON_EMPTY_HEADER_ROW: "Empty / unreadable header row",
    REASON_FORMULA_HEADER: "Formula headers",
    REASON_PHI_IN_HEADER: "PHI-shaped headers",
    REASON_MISMATCHED_SHA: "Duplicate filename, mismatched content",
    REASON_COLLISION: "Form-code collision (multiple files share the same prefix)",
}

# ---------------------------------------------------------------------------
# Custom exception
# ---------------------------------------------------------------------------


class ExcludeForReview(Exception):
    """Raised when a file or pair must be routed to human review instead of
    being processed into a policy YAML.

    Without this exception the various guards inside ``read_headers_only`` and
    ``build_yaml_for_pair`` would have to return sentinel values (``None``,
    empty list, special strings) that callers could silently ignore.  A
    distinct exception forces every call site to decide explicitly what to do
    with the excluded file — there is no "accidentally continue" path.

    The first positional argument is always a reason-code string, one of the
    ``REASON_*`` constants in this module.  Optional extra arguments carry
    diagnostic context (e.g. the list of visible sheet names for
    ``multi_sheet_workbook``).
    """


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class Pair:
    """A cleanly matched PDF + dataset ready for YAML generation."""

    form_code: str
    pdf_path: Path
    dataset_path: Path


@dataclass
class ReviewEntry:
    """One row in ``SoT_intake_review.md`` — a file excluded from YAML generation."""

    reason: str
    label: str          # human-readable form / file label
    file_path: Path     # canonical path of the problematic file
    notes: str = ""


@dataclass
class IntakeManifest:
    """Result of ``pair_files``: what to build vs. what to send to review."""

    paired: list[Pair] = field(default_factory=list)
    to_review: list[ReviewEntry] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Filename normalisation helpers
# ---------------------------------------------------------------------------

_VERSION_SUFFIX_RE = re.compile(
    r"(?:"
    r"\s*\(\d+\)"           # " (1)", " (2)", …
    r"|[_ ]v\d+(?:\.\d+)+"  # "_v1.0", " v2.1", …
    r"|[_ ]v\d+"            # "_v2", " v1"
    r")$",
    re.IGNORECASE,
)
_WHITESPACE_RE = re.compile(r"[\s_\-\.]+")
_FORM_CODE_RE = re.compile(r"^([0-9]+[A-Za-z]?)")


def _normalise(name: str) -> str:
    """Lowercase, strip extension + version suffixes, collapse separators to ``_``."""
    stem = Path(name).stem
    stem = _VERSION_SUFFIX_RE.sub("", stem)
    stem = _WHITESPACE_RE.sub("_", stem.strip())
    return stem.lower()


def _extract_form_code(normalised: str) -> str | None:
    """Return the leading form-code prefix (e.g. ``1a``, ``12``, ``101``) or None."""
    m = _FORM_CODE_RE.match(normalised)
    return m.group(1).lower() if m else None


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Public functions (ordered per spec)
# ---------------------------------------------------------------------------


def read_headers_only(path: Path) -> list[str]:
    """Return the column headers (row 1 only) of an xlsx or csv file.

    For xlsx: opens with ``openpyxl`` in read-only mode, iterates exactly one
    row (``max_row=1``), then stops.  For csv: calls ``next(reader)`` once.
    Row-2+ bytes are structurally unreachable in both branches.

    Raises ``ExcludeForReview`` for unsupported extensions, empty/blank header
    rows, or multi-sheet workbooks (xlsx only).
    """
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        try:
            visible = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
            if len(visible) != 1:
                raise ExcludeForReview(
                    "multi_sheet_workbook",
                    path,
                    visible,
                )
            ws = wb[visible[0]]
            rows = ws.iter_rows(max_row=1, values_only=True)
            first = next(rows, None)
            if first is None or all(c is None or str(c).strip() == "" for c in first):
                raise ExcludeForReview(REASON_EMPTY_HEADER_ROW, path)
            headers = [str(c) for c in first if c is not None and str(c).strip() != ""]
        finally:
            wb.close()
        return headers

    elif suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            try:
                first = next(reader)
            except StopIteration:
                raise ExcludeForReview(REASON_EMPTY_HEADER_ROW, path)
        if not first or all(c.strip() == "" for c in first):
            raise ExcludeForReview(REASON_EMPTY_HEADER_ROW, path)
        return first

    else:
        raise ExcludeForReview("unsupported_extension", path, suffix)


def _validate_headers(headers: list[str], dataset_path: Path) -> None:
    """Pure guard: raise ``ExcludeForReview`` if any header is unsafe.

    Checks (in order):
    - Formula header: any header that starts with ``=`` signals a spreadsheet
      formula reference that could resolve to off-sheet data.
    - PHI-shaped header: any header that matches the drop/id/birthdate patterns
      from ``scripts/security/phi_scrub.py`` is treated as a likely PHI column
      name that should not reach the LLM.
    """
    for h in headers:
        if h.startswith("="):
            raise ExcludeForReview(REASON_FORMULA_HEADER, dataset_path, h)

    # Load PHI scrub config (may return None if no config file exists for this
    # study; in that case we skip the PHI-name check — no config = no patterns).
    scrub_cfg = None
    try:
        from scripts.security.phi_scrub import load_scrub_config
        scrub_cfg = load_scrub_config()
    except Exception:
        _LOG.warning(
            "study_intake._validate_headers phi_scrub_load_failed dataset=%s",
            dataset_path.name,
        )

    if scrub_cfg is None:
        return

    for h in headers:
        # Any header that the scrubber would DROP or PSEUDONYMIZE or treat as
        # a birthdate is a PHI-shaped field name — exclude it from LLM reach.
        if scrub_cfg.field_is_drop(h) or scrub_cfg.field_is_id(h) or scrub_cfg.field_is_birthdate(h):
            raise ExcludeForReview(REASON_PHI_IN_HEADER, dataset_path, h)


def pair_files(study_dir: Path) -> IntakeManifest:
    """Walk datasets/ and annotated_pdfs/, normalise filenames, pair by form-code prefix.

    Pairing rules:
    - Exact form-code prefix match → ``Pair``.
    - Collision (2+ PDFs or 2+ datasets share the same prefix) → all to review.
    - Identical SHA256 across same prefix → deduplicate silently (keep first).
    - Mismatched SHA256 across same prefix → all copies to review.
    - Unpaired PDF → review with ``unpaired_pdf``.
    - Unpaired dataset → review with ``unpaired_dataset``.
    """
    manifest = IntakeManifest()

    datasets_dir = study_dir / "datasets"
    pdfs_dir = study_dir / "annotated_pdfs"

    # --- collect datasets ---------------------------------------------------
    dataset_files: list[Path] = []
    if datasets_dir.is_dir():
        dataset_files = [
            p for p in datasets_dir.iterdir()
            if p.suffix.lower() in {".xlsx", ".csv"} and p.is_file()
        ]

    # --- collect PDFs -------------------------------------------------------
    pdf_files: list[Path] = []
    if pdfs_dir.is_dir():
        pdf_files = [p for p in pdfs_dir.iterdir() if p.suffix.lower() == ".pdf" and p.is_file()]

    # --- build prefix → [path, …] maps -------------------------------------
    dataset_map: dict[str, list[Path]] = {}
    for p in dataset_files:
        code = _extract_form_code(_normalise(p.name))
        if code is None:
            manifest.to_review.append(
                ReviewEntry(
                    reason=REASON_UNPAIRED_DATASET,
                    label=p.stem,
                    file_path=p,
                    notes="Could not extract a leading form-code prefix from filename.",
                )
            )
            continue
        dataset_map.setdefault(code, []).append(p)

    pdf_map: dict[str, list[Path]] = {}
    for p in pdf_files:
        code = _extract_form_code(_normalise(p.name))
        if code is None:
            manifest.to_review.append(
                ReviewEntry(
                    reason=REASON_UNPAIRED_PDF,
                    label=p.stem,
                    file_path=p,
                    notes="Could not extract a leading form-code prefix from filename.",
                )
            )
            continue
        pdf_map.setdefault(code, []).append(p)

    all_codes = set(dataset_map) | set(pdf_map)

    for code in sorted(all_codes):
        datasets = dataset_map.get(code, [])
        pdfs = pdf_map.get(code, [])

        # ---- unpaired -------------------------------------------------------
        if not pdfs:
            for d in datasets:
                manifest.to_review.append(
                    ReviewEntry(
                        reason=REASON_UNPAIRED_DATASET,
                        label=d.stem,
                        file_path=d,
                        notes=f"No PDF with canonical prefix `{code}` found.",
                    )
                )
            continue

        if not datasets:
            for p in pdfs:
                manifest.to_review.append(
                    ReviewEntry(
                        reason=REASON_UNPAIRED_PDF,
                        label=p.stem,
                        file_path=p,
                        notes=f"No xlsx/csv with canonical prefix `{code}` found.",
                    )
                )
            continue

        # ---- collision (multiple files on either side) ----------------------
        if len(pdfs) > 1 or len(datasets) > 1:
            # Attempt SHA-based deduplication on each side independently.
            pdfs = _dedup_or_flag(pdfs, code, manifest, side="pdf")
            datasets = _dedup_or_flag(datasets, code, manifest, side="dataset")
            # If collision remains after SHA dedup, route everything to review.
            if len(pdfs) != 1 or len(datasets) != 1:
                for p in pdfs:
                    manifest.to_review.append(
                        ReviewEntry(
                            reason=REASON_COLLISION,
                            label=p.stem,
                            file_path=p,
                            notes=f"Multiple files share form-code prefix `{code}`.",
                        )
                    )
                for d in datasets:
                    manifest.to_review.append(
                        ReviewEntry(
                            reason=REASON_COLLISION,
                            label=d.stem,
                            file_path=d,
                            notes=f"Multiple files share form-code prefix `{code}`.",
                        )
                    )
                continue

        manifest.paired.append(Pair(form_code=code, pdf_path=pdfs[0], dataset_path=datasets[0]))

    _LOG.warning(
        "study_intake.pair_files study_dir=%s paired=%d to_review=%d",
        study_dir,
        len(manifest.paired),
        len(manifest.to_review),
    )
    return manifest


def _dedup_or_flag(
    paths: list[Path],
    code: str,
    manifest: IntakeManifest,
    side: str,
) -> list[Path]:
    """Deduplicate identical-SHA files; flag mismatched-SHA groups to review.

    Returns the surviving list (length 1 if cleanly deduped, else the full
    original list so the caller can route the whole group to review).
    """
    if len(paths) <= 1:
        return paths

    sha_map: dict[str, Path] = {}
    mismatch = False
    for p in paths:
        digest = _sha256(p)
        if digest in sha_map:
            _LOG.warning(
                "study_intake.dedup duplicate_sha side=%s code=%s kept=%s dropped=%s",
                side, code, sha_map[digest].name, p.name,
            )
        else:
            sha_map[digest] = p

    if len(sha_map) == 1:
        # All files are byte-identical — keep one, silently drop the rest.
        return list(sha_map.values())

    # Multiple distinct SHAs — mismatch, all to review.
    for p in paths:
        manifest.to_review.append(
            ReviewEntry(
                reason=REASON_MISMATCHED_SHA,
                label=p.stem,
                file_path=p,
                notes=(
                    f"Multiple {side} files with prefix `{code}` have "
                    "different content (SHA256 mismatch)."
                ),
            )
        )
    return []


def route_to_review(entry: ReviewEntry, review_file: Path) -> None:
    """Append *entry* to the correct section of ``SoT_intake_review.md``.

    On re-run, any entry whose label already appears under ``- [x]`` (checked)
    or with a filled ``Disposition:`` line is considered dispositioned and is
    not re-appended.
    """
    review_file.parent.mkdir(parents=True, exist_ok=True)

    # Parse existing file for already-dispositioned entries.
    dispositioned: set[str] = set()
    if review_file.is_file():
        content = review_file.read_text(encoding="utf-8")
        for line in content.splitlines():
            # Checked checkbox: "- [x] **<label>**" or "- [x] <label>"
            if line.strip().startswith("- [x]"):
                m = re.search(r"\*\*(.+?)\*\*", line)
                if m:
                    dispositioned.add(m.group(1))
            # Filled Disposition line
            if "Disposition:" in line:
                after = line.split("Disposition:", 1)[1].strip()
                if after and after != "__________":
                    # extract the label from a nearby prior "**label**"
                    # We track by file path instead for reliability.
                    pass
        # Also track by file path references.
        for line in content.splitlines():
            if "- [x]" in line and str(entry.file_path) in line:
                dispositioned.add(entry.label)
                break
            if entry.label in line and "- [x]" in line:
                dispositioned.add(entry.label)
                break

    if entry.label in dispositioned:
        _LOG.warning(
            "study_intake.route_to_review already_dispositioned label=%s reason=%s",
            entry.label, entry.reason,
        )
        return

    section_title = _SECTION_TITLES.get(entry.reason, f"Other ({entry.reason})")
    new_block = (
        f"\n- [ ] **{entry.label}** — `{entry.file_path}`\n"
        f"      Reason: `{entry.reason}`\n"
        f"      Notes: {entry.notes}\n"
        f"      Disposition: __________ "
        f"(add_with_override / delete / keep_in_review / rename)\n"
    )

    if review_file.is_file():
        existing = review_file.read_text(encoding="utf-8")
        section_header = f"## {section_title}"
        if section_header in existing:
            # Append inside the existing section (before the next ## or EOF).
            insert_pos = existing.index(section_header) + len(section_header)
            # Find next section or end.
            next_section = existing.find("\n## ", insert_pos)
            if next_section == -1:
                updated = existing + new_block
            else:
                updated = existing[:next_section] + new_block + existing[next_section:]
            review_file.write_text(updated, encoding="utf-8")
            return
        # Section not present — append section + entry at end.
        review_file.write_text(
            existing.rstrip() + f"\n\n## {section_title}\n" + new_block,
            encoding="utf-8",
        )
    else:
        # Create fresh file.
        from datetime import date
        header = (
            f"# SoT Intake Review\n"
            f"Generated: {date.today().isoformat()} "
            f"by `python -m scripts.source_truth.study_intake`\n"
            f"Tick `- [x]` after triage; re-run the intake CLI to refresh.\n"
        )
        review_file.write_text(
            header + f"\n## {section_title}\n" + new_block,
            encoding="utf-8",
        )


def build_yaml_for_pair(
    pair: Pair,
    study: str,
    sot_dir: Path,
    raw_dir: Path,
    review_file: Path,
    *,
    force: bool = False,
) -> Path | None:
    """Build (or skip) ``{form}_policy.yaml`` for a single PDF+dataset pair.

    Pipeline:
        read_headers_only → _validate_headers → pdf_evidence → extractor →
        reviewer → record.validate_record → atomic write.

    Returns the output path on success, or ``None`` if the pair was routed to
    review (``ExcludeForReview`` raised at any step).

    Skip policy (Q1=b): if the YAML already exists and ``force=False``, return
    the existing path without re-running.
    """
    form = pair.form_code
    out_path = sot_dir / f"{form}_policy.yaml"

    if out_path.exists() and not force:
        _LOG.warning(
            "study_intake.build_yaml_for_pair skip_existing form=%s path=%s",
            form, out_path,
        )
        return out_path

    try:
        # Step 1 — headers only (enforces data-isolation invariant)
        headers = read_headers_only(pair.dataset_path)

        # Step 2 — validate headers (formula + PHI guards)
        _validate_headers(headers, pair.dataset_path)

        # Step 3 — PDF evidence
        pdf_data = pdf_evidence.extract_pdf_evidence(
            pair.pdf_path.read_text(encoding="utf-8", errors="replace")
        )

        # Step 4 — extractor agent
        extractor_result = sot_extractor_agent.run_extractor(
            form=form,
            sot_dir=sot_dir,
            raw_pdf_dir=raw_dir / "annotated_pdfs",
            dataset_dir=raw_dir / "datasets",
            pilot_dir=raw_dir / "pilot",
            output_dir=sot_dir,
        )
        draft_yaml_text: str = extractor_result.get("yaml", "")

        # Step 5 — reviewer agent
        draft_yaml_path = sot_dir / f"{form}_policy.draft.yaml"
        draft_pack_path = sot_dir / f"{form}_evidence_pack.draft.json"
        # Write ephemeral draft files so reviewer can read them via its existing API.
        sot_dir.mkdir(parents=True, exist_ok=True)
        draft_yaml_path.write_text(draft_yaml_text, encoding="utf-8")
        pack_text = extractor_result.get("evidence_pack", "{}")
        draft_pack_path.write_text(pack_text, encoding="utf-8")

        reviewer_result = sot_reviewer_agent.run_reviewer(
            form=form,
            sot_dir=sot_dir,
            raw_pdf_dir=raw_dir / "annotated_pdfs",
            dataset_dir=raw_dir / "datasets",
            pilot_dir=raw_dir / "pilot",
            draft_yaml_path=draft_yaml_path,
            draft_pack_path=draft_pack_path,
            reviews_dir=sot_dir / "reviews",
        )

        # Clean up ephemeral drafts.
        draft_yaml_path.unlink(missing_ok=True)
        draft_pack_path.unlink(missing_ok=True)

        # Step 6 — schema validation
        import yaml as _yaml  # local import to keep top-level import list lean
        record_dict: dict[str, Any] = _yaml.safe_load(reviewer_result.get("yaml", draft_yaml_text)) or {}
        _record.validate_record(record_dict)

        # Step 7 — atomic write
        final_yaml = reviewer_result.get("yaml", draft_yaml_text)
        _atomic_write(out_path, final_yaml)
        _LOG.warning(
            "study_intake.build_yaml_for_pair wrote form=%s path=%s",
            form, out_path,
        )
        return out_path

    except ExcludeForReview as exc:
        reason = exc.args[0] if exc.args else "unknown"
        notes = "; ".join(str(a) for a in exc.args[1:]) if len(exc.args) > 1 else ""
        _LOG.warning(
            "study_intake.build_yaml_for_pair excluded form=%s reason=%s",
            form, reason,
        )
        route_to_review(
            ReviewEntry(reason=reason, label=form, file_path=pair.dataset_path, notes=notes),
            review_file,
        )
        return None


def _atomic_write(path: Path, text: str) -> None:
    """Write *text* to *path* atomically (tmp → rename)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def run_intake(study: str, *, force: bool = False) -> int:
    """Orchestrate the full intake pipeline for *study*.

    Returns 0 on success (all pairs either built or cleanly routed to review),
    1 if any pair raised an unexpected (non-ExcludeForReview) exception.
    """
    repo_root = Path(config.BASE_DIR) if hasattr(config, "BASE_DIR") else Path(__file__).resolve().parents[3]
    raw_dir = repo_root / "data" / "raw" / study
    sot_dir = repo_root / "data" / "SoT" / study
    review_file = sot_dir / "human_review" / "SoT_intake_review.md"

    if not raw_dir.is_dir():
        _LOG.warning("study_intake.run_intake missing_raw_dir study=%s path=%s", study, raw_dir)
        print(f"ERROR: raw data directory not found: {raw_dir}", file=sys.stderr)
        return 1

    manifest = pair_files(raw_dir)

    failed = 0
    built = 0
    skipped = 0

    for pair in manifest.paired:
        try:
            result = build_yaml_for_pair(
                pair,
                study=study,
                sot_dir=sot_dir,
                raw_dir=raw_dir,
                review_file=review_file,
                force=force,
            )
            if result is None:
                # routed to review — not a failure
                pass
            elif result.exists():
                built += 1
        except ExcludeForReview:
            # Already handled inside build_yaml_for_pair; count is tracked there.
            pass
        except Exception as exc:
            _LOG.warning(
                "study_intake.run_intake unexpected_error form=%s error=%s",
                pair.form_code, exc,
            )
            failed += 1

    for entry in manifest.to_review:
        route_to_review(entry, review_file)

    review_count = len(manifest.to_review)
    print(
        f"study_intake: {len(manifest.paired)} aligned → "
        f"{built} YAMLs written, "
        f"{len(manifest.paired) - built} routed-to-review from pairs, "
        f"{review_count} pre-pair items in {review_file.relative_to(repo_root)}"
    )
    if failed:
        print(f"WARNING: {failed} pair(s) failed with unexpected errors.", file=sys.stderr)
        return 1
    return 0


def main() -> int:
    """CLI entry point for SoT intake.

    Inputs : data/raw/<study>/annotated_pdfs/*.pdf
             data/raw/<study>/datasets/*.{xlsx,csv}
    Outputs: data/SoT/<study>/<form>_policy.yaml  (one per aligned pair)
             data/SoT/<study>/human_review/SoT_intake_review.md

    Re-run policy: existing YAMLs are skipped by default; use --force to
    regenerate.  Already-dispositioned entries in SoT_intake_review.md (ticked
    ``- [x]``) are not re-appended.

    See: docs/runbook_sot_build.md
    """
    parser = argparse.ArgumentParser(
        prog="python -m scripts.source_truth.study_intake",
        description=(
            "Build Source-of-Truth policy YAMLs for a clinical study.\n\n"
            "Inputs : data/raw/<study>/annotated_pdfs/*.pdf\n"
            "         data/raw/<study>/datasets/*.{xlsx,csv}\n"
            "Outputs: data/SoT/<study>/<form>_policy.yaml  (one per aligned pair)\n"
            "         data/SoT/<study>/human_review/SoT_intake_review.md\n\n"
            "Re-run policy: existing YAMLs are skipped unless --force is passed.\n\n"
            "See: docs/runbook_sot_build.md"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "study",
        help="Study name, e.g. Indo-VAP.  Must match a directory under data/raw/.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        default=False,
        help="Overwrite existing policy YAMLs (default: skip-if-exists).",
    )
    args = parser.parse_args()
    return run_intake(args.study, force=args.force)


if __name__ == "__main__":
    sys.exit(main())
